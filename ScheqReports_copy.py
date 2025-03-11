import pymongo
import pandas as pd
import re
import os
import smtplib
import tempfile
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image

# MongoDB Connection
MONGO_URI = "mongodb://192.168.0.34:27017/"
DATABASE_NAME = "scheq"
client = pymongo.MongoClient(MONGO_URI)
db = client[DATABASE_NAME]

# Get the current month and year
current_month = datetime.now().strftime('%Y-%m')
start_of_month = datetime(datetime.now().year, datetime.now().month, 1)
end_of_month = datetime(datetime.now().year, datetime.now().month + 1, 1) if datetime.now().month < 12 else datetime(datetime.now().year + 1, 1, 1)

# Read CSV for page-sheet mapping
daily_checksheet_df = pd.read_csv(r'daily_checksheet.csv')

# Get collections ending with "ds"
collections = [col for col in db.list_collection_names() if re.search(r'ds$', col)]

current_year = datetime.now().year
current_month = datetime.now().month
current_day = datetime .now().day
# Temporary file path
temp_dir = ''
excel_filename = os.path.join(temp_dir, f"PM_Module_{current_day}_{current_month}_{current_year}.xlsx")
sheet_written = False

# Dictionary for unique sheet names
collection_sheet_map = {}

# Define thick border style
thick_border = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

# Define thin border style for date columns
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Assign unique sheet names
for collection_name in collections:
    row = daily_checksheet_df[daily_checksheet_df['page name'] == collection_name]
    sheet_name = row.iloc[0, daily_checksheet_df.columns.get_loc('page name') + 1] if not row.empty else collection_name
    count = 1
    while sheet_name in collection_sheet_map.values():
        sheet_name = f"{sheet_name}_{count}"
        count += 1
    collection_sheet_map[collection_name] = sheet_name

# Create and write Excel file
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    for collection_name, sheet_name in collection_sheet_map.items():
        collection = db[collection_name]
        data = list(collection.find({'submissionDate': {'$gte': start_of_month, '$lt': end_of_month}}))

        if data:
            sheet_written = True
            df = pd.DataFrame(data)
            df.drop(columns=['_id', '__v'], errors='ignore', inplace=True)

            cols = ['opName'] + [col for col in df.columns if col != 'opName']
            df = df[cols]  # Reorder DataFrame columns

            print("Columns after reordering:", df.columns.tolist())  # Verify new order

            # Transpose the DataFrame
            if 'submissionDate' in df.columns:
                df['submissionDate'] = pd.to_datetime(df['submissionDate']).dt.date
                df.set_index('submissionDate', inplace=True)

            df = df.T  # Transposing data

            df.to_excel(writer, sheet_name=sheet_name, index=True, startrow=4, startcol=3)

            # Apply styling
            sheet = writer.sheets[sheet_name]

            # Enable text wrapping and adjust row height
            for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                sheet.row_dimensions[row[0].row].height = 30  # Set a default row height

            # Adjust column width
            sheet.column_dimensions['A'].width = 5  # Reduced width of serial number column
            sheet.column_dimensions['B'].width = 30  # Increased width for Check Points column
            sheet.column_dimensions['C'].width = 30  # Increased width for Acceptance Criteria column
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                if column not in ['A', 'B', 'C']:  # Skip columns A, B, and C
                    for cell in col:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    sheet.column_dimensions[column].width = min(max_length + 20, 50)  # Limit column width

            # Merge cells and add headings
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
            heading_cell = sheet.cell(row=1, column=1)
            heading_cell.value = "Daily Checksheet Report - VT1014532-F10"
            heading_cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            heading_cell.font = Font(bold=True, size=20, color="FFFFFF")
            heading_cell.alignment = Alignment(horizontal="center", vertical="center")

            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=sheet.max_column)
            month_title_cell = sheet.cell(row=2, column=1)
            month_title_cell.value = f"Report for {datetime.now().strftime('%B %Y')}"
            month_title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            month_title_cell.alignment = Alignment(horizontal="center", vertical="center")
            month_title_cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")

            # Split sheet_name by "|" and place in two different cells above the "Check Points" field
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
            sheet_name_parts = sheet_name.split("|")
            sheet.cell(row=3, column=1, value=sheet_name_parts[0].strip())  # Moved to left side
            sheet.cell(row=3, column=1).font = Font(bold=True, size=12)
            sheet.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")

            if len(sheet_name_parts) > 1:
                sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
                sheet.cell(row=4, column=1, value=sheet_name_parts[1].strip())  # Moved to left side
                sheet.cell(row=4, column=1).font = Font(bold=True, size=12)
                sheet.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center")
            
            # Add "Department: UTM & EHS" text
            sheet.cell(row=3, column=3, value="Department: UTM & EHS")
            sheet.cell(row=3, column=3).font = Font(bold=True, size=12)
            sheet.cell(row=3, column=3).alignment = Alignment(horizontal="left", vertical="center")

            # Add a new column for serial numbers to the left of "Check Points"
            sheet.cell(row=5, column=1, value="Sr.No.")  # Serial number heading
            sheet.cell(row=5, column=1).font = Font(bold=True, size=15)
            # Remove orange fill and use white background
            sheet.cell(row=5, column=1).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            sheet.cell(row=5, column=1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            sheet.cell(row=5, column=1).border = thick_border

            # Increase the height of the row containing the fields
            sheet.row_dimensions[5].height = 40  # Increased row height

            green_fill = PatternFill(start_color="28a744", end_color="28a744", fill_type="solid")
            # Removing orange fill and using white background
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_font = Font(bold=True, size=12)

            # Insert values into the second column with heading "Check Points"
            sheet.cell(row=5, column=2, value="Check Points")  # Set heading
            sheet.cell(row=5, column=2).font = Font(bold=True, size=15)
            sheet.cell(row=5, column=2).fill = white_fill  # Changed from orange to white
            sheet.cell(row=5, column=2).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            sheet.cell(row=5, column=2).border = thick_border

            # Insert values into the third column with heading "Acceptance Criteria"
            sheet.cell(row=5, column=3, value="Acceptance Criteria")  # Set heading
            sheet.cell(row=5, column=3).font = Font(bold=True, size=15)
            sheet.cell(row=5, column=3).fill = white_fill  # Changed from orange to white
            sheet.cell(row=5, column=3).border = thick_border
            sheet.cell(row=5, column=3).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Insert values into the fourth column with heading "Inspection Methods"
            sheet.cell(row=5, column=4, value="Inspection Methods")  # Set heading
            sheet.cell(row=5, column=4).font = Font(bold=True, size=15)
            sheet.cell(row=5, column=4).fill = white_fill  # Changed from orange to white
            sheet.cell(row=5, column=4).border = thick_border
            sheet.cell(row=5, column=4).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Extract values from the next-to-next column
            row = daily_checksheet_df[daily_checksheet_df['page name'] == collection_name]
            if not row.empty:
                next_to_next_col_index = daily_checksheet_df.columns.get_loc('page name') + 2
                values = row.iloc[0, next_to_next_col_index].split('|') if pd.notna(row.iloc[0, next_to_next_col_index]) else []

            # Fill values into cells
            for i, value in enumerate(values):
                # Remove text between "-" from the second column
                cleaned_value = re.sub(r'-.*?-', '', value).strip()
                sheet.cell(row=6 + i, column=2, value=cleaned_value)
                sheet.cell(row=6 + i, column=2).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                sheet.cell(row=6 + i, column=2).border = thick_border

                # Extract text between "-" characters for the third column
                match = re.search(r'-(.*?)-', value)
                if match:
                    acceptance_criteria = match.group(1).strip()
                    sheet.cell(row=6 + i, column=3, value=acceptance_criteria)
                    sheet.cell(row=6 + i, column=3).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                    sheet.cell(row=6 + i, column=3).border = thick_border

                # Add "Visual Inspection & Record" to the fourth column
                sheet.cell(row=6 + i, column=4, value="Visual Inspection & Record")
                sheet.cell(row=6 + i, column=4).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                sheet.cell(row=6 + i, column=4).border = thick_border

                # Add serial number for each row
                sheet.cell(row=6 + i, column=1, value=i + 1)  # Serial numbers start from 1
                sheet.cell(row=6 + i, column=1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                sheet.cell(row=6 + i, column=1).border = thick_border

            # Format date columns (columns after the "Inspection Methods" column)
            for col in range(5, sheet.max_column + 1):  # Starting from column 5 (after "Inspection Methods")
                # Format header (date) cell
                date_header_cell = sheet.cell(row=5, column=col)
                date_header_cell.font = Font(bold=True, size=15)  # Increase font size for dates
                date_header_cell.fill = white_fill  # Changed from orange to white
                date_header_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                date_header_cell.border = thin_border  # Add thin border for date headers
                
                # Format all cells in the date column
                for row in range(6, sheet.max_row + 1):
                    date_cell = sheet.cell(row=row, column=col)
                    date_cell.font = Font(size=12)  # Adjust size as needed
                    date_cell.alignment = Alignment(vertical="center", horizontal="center")
                    date_cell.border = thin_border  # Add thin border for date cells

            # Apply thick borders only to the first 4 columns (S.No., Check Points, Acceptance Criteria, Inspection Methods)
            for row in range(5, sheet.max_row + 1):
                for col in range(1, 5):  # Only apply to columns 1-4
                    sheet.cell(row=row, column=col).border = thick_border

            print(f"Exported {collection_name} to sheet {sheet_name}")
        else:
            print(f"Skipping {collection_name} (No data found)")


client.close()
