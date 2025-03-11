import pymongo
import pandas as pd
import re
import os
import smtplib
import tempfile
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import io
import numpy as np

# MongoDB Connection
MONGO_URI = "mongodb://192.168.0.34:27017/"
DATABASE_NAME = "scheq"
client = pymongo.MongoClient(MONGO_URI)
db = client[DATABASE_NAME]

# Get current month and year
current_year = datetime.now().year
current_month = datetime.now().month
current_day = datetime .now().day
month_name = datetime.now().strftime('%B').upper()
num_days = (datetime(current_year, current_month + 1, 1) - datetime(current_year, current_month, 1)).days
date_headers = [(datetime(current_year, current_month, i).strftime('%d-%b-%Y')) for i in range(1, num_days + 1)]

# Read CSV for page-sheet mapping
daily_checksheet_df = pd.read_csv(r'daily_checksheet.csv')

# Filter collections that exist in MongoDB and are listed in the CSV
db_collections = db.list_collection_names()
valid_collections = daily_checksheet_df[daily_checksheet_df['page name'].isin(db_collections)]

# Create a temporary Excel file
temp_dir = ''
excel_filename = os.path.join(temp_dir, f"PM_MAINTENANCE_{current_day}_{current_month}_{current_year}.xlsx")

# Create an Excel workbook
wb = Workbook()
sheet = wb.active
sheet.title = "PM Maintenance"

# Define headers
header = ["Sl No.", "Machine / Station Name", "ID. NO.", "Category", "Frequency"] + date_headers

# Merge and format title
sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(header))
title_cell = sheet.cell(row=1, column=1)
title_cell.value = f"DAILY PREDICTIVE AND PREVENTIVE REPORT - {month_name} {current_year}"
title_cell.font = Font(bold=True, size=16)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
sheet.row_dimensions[1].height = 30  # Increase row height for the title

# Define headers with dark green background
header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
thick_border = Border(left=Side(style="medium", color="90EE90"),
                      right=Side(style="medium", color="90EE90"),
                    #   top=Side(style="medium", color="90EE90"),
                    #   bottom=Side(style="medium", color="90EE90")
                    )
light_border = Border(left=Side(style="thin", color="90EE90"),
                      right=Side(style="thin", color="90EE90"),
                    #   top=Side(style="thin", color="90EE90"),
                    #   bottom=Side(style="thin", color="90EE90")
                    )

for col_num, column_title in enumerate(header, 1):
    cell = sheet.cell(row=3, column=col_num, value=column_title)
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_num <= 5:  # Apply thick borders for important columns
        cell.border = thick_border
    else:
        cell.border = light_border
    if col_num in [1, 4, 5]:  # Reduce width for Sl No, ID No, Category, and Frequency
        sheet.column_dimensions[cell.column_letter].width = 10
    elif col_num == 2:  # Adjust width for Machine Name
        sheet.column_dimensions[cell.column_letter].width = 50  # Increase this value to make it wider
    elif col_num == 3:  # Adjust width for ID No.
        sheet.column_dimensions[cell.column_letter].width = 30  # Increase this value as needed
    else:  # Reduce width for date columns
        sheet.column_dimensions[cell.column_letter].width = 12

# Increase the height of the header row (row 3)
sheet.row_dimensions[3].height = 25  # Set header row height to 25

# Fetch and insert data
sl_no = 1
data_rows = []
for _, row in valid_collections.iterrows():
    collection_name = row['page name']
    extracted_text = row['extracted <h1> text'] if pd.notna(row['extracted <h1> text']) else "N/A|N/A"
    machine_name, id_no = extracted_text.split('|') if '|' in extracted_text else (extracted_text, "N/A")
    
    collection = db[collection_name]
    data = list(collection.find({}))  # Fetch all data
    print(f"Processing collection: {collection_name}, records found: {len(data)}")
    
    # Initialize row with default values
    row_data = [sl_no, machine_name.strip(), id_no.strip(), "C", "Daily"] + ["" for _ in range(num_days)]
    
    # Fill 'P' based on submissionDate field
    for record in data:
        if "submissionDate" in record:
            submission_date = record["submissionDate"]
            if isinstance(submission_date, dict) and "$date" in submission_date:
                submission_date = datetime.strptime(submission_date["$date"][:10], "%Y-%m-%d")
            if isinstance(submission_date, datetime) and submission_date.year == current_year and submission_date.month == current_month:
                day_index = submission_date.day - 1  # Convert to zero-based index
                row_data[5 + day_index] = "P"
    
    data_rows.append(row_data)
    sheet.append(row_data)
    sl_no += 1

# Apply borders and alignment to all data cells
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=len(header)):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = light_border

# Apply thick borders to key columns
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.border = thick_border

# Align Machine Name to the left
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Increase row height for data rows to make cells more spacious
for row in range(4, sheet.max_row + 1):
    sheet.row_dimensions[row].height = 20  # Set row height to 20 (default is 15)

# Generate the graph
total_rows = len(data_rows)
p_counts = [0] * num_days

for row in data_rows:
    for i in range(num_days):
        if row[5 + i] == "P":
            p_counts[i] += 1

dates = [datetime(current_year, current_month, i + 1).strftime('%d-%b') for i in range(num_days)]

# Create a grouped bar plot
x = np.arange(len(dates))  # the label locations
width = 0.3  # the width of the bars (reduced to add gap)

fig, ax = plt.subplots(figsize=(15, 3))  # Increase figure size for better visibility
rects1 = ax.bar(x - width/1.5, [total_rows] * num_days, width, label='Assigned', color='green')  # Left bar: Assigned (green)
rects2 = ax.bar(x + width/1.5, p_counts, width, label='Actual', color='grey')  # Right bar: Actual (grey)

# Add labels, title, and legend
ax.set_xlabel('Date')
ax.set_ylabel('Count')
ax.set_title('Daily Maintenance Report')
ax.set_xticks(x)
ax.set_xticklabels(dates, rotation=45, ha='right')  # Rotate x-axis labels for better visibility
ax.legend()

# Ensure y-axis labels are whole numbers
ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))

# Save the plot to a BytesIO object
buf = io.BytesIO()
plt.savefig(buf, format='png', bbox_inches='tight')  # Use bbox_inches='tight' to avoid cutting off labels
buf.seek(0)
img = Image(buf)

# Insert the image below the table
sheet.add_image(img, f'A{sheet.max_row + 2}')

# Save the Excel file
wb.save(excel_filename)


client.close()