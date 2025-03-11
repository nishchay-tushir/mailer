
import pymongo
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def fetch_monthly_kwh_for_meters(start_time, end_time, meter_categories, exclude_from_total):

    meters_to_divide = {
        "Pump Room main meter", "Mechanical AHU-01", "Mechanical AHU-02", "Mechanical AHU-03",
        "Mechanical AHU-04", "Technical Penthouse -L-2 Main meter", "Electronics AHU-1",
        "Electronics AHU-2", "Electronics AHU-4", "Electronics AHU-5", "Electronics AHU-6",
        "Electronics AHU-7", "PCWP-Mechanical", "Air cool chiller Mechanical (CH4)",
        "Air cool chiller Electronics (CH-3)", "CLWS-2 (Mechanical)", "Air Compressor Main meter",
        "Technical Penthouse-L-1 Incomer main panel", "Electronics UPS Incomer",
        "Electronics UPS Outgoing - 1", "Electronics UPS Outgoing - 3", "BMS Room ELDB",
        "BMS Room-Aux DB - 2", "BMS Room- MLDB (Electronics DB)", "Electronics AHU main panel",
        "Canteen main panel", "Transformer No. 1", "Transformer No. 2", "Transformer No. 3",
        "Transformer No. 4", "Incoming 500 KVA DG -6", "Panel Room-Mech & WH ELDB",
        "Panel Room-Aux DB-1","Panel Room- LDB-Mech,Pump Room, security, WRC AHU","BMS Room- MLDB ( Electronics DB )"
    }

    client = pymongo.MongoClient("mongodb://192.168.0.34:27017/")
    db = client["edms"]
    collection = db["kwh"]

    all_data = []

    for category, meters in meter_categories.items():
        category_data = []

        for meter_name in meters:
            cursor = collection.find({
                "timestamp": {"$gte": start_time.strftime("%Y-%m-%d"), "$lt": end_time.strftime("%Y-%m-%d")}
            })

            data = {"Timestamp": [], "KWH_VALUES": []}
            for doc in cursor:
                timestamp = pd.to_datetime(doc["timestamp"])
                if meter_name in doc["device_names"]:
                    index = doc["device_names"].index(meter_name)
                    data["Timestamp"].append(timestamp)
                    data["KWH_VALUES"].append(doc["kwh_values"][index])

            if not data["Timestamp"]:
                continue

            df = pd.DataFrame(data).sort_values(by="Timestamp").set_index("Timestamp")
            df["KWH_DIFF"] = df["KWH_VALUES"].diff().fillna(0)

            # Divide KWH values by 1000 for specific meters
            if meter_name in meters_to_divide:
                df["KWH_DIFF"] = df["KWH_DIFF"] / 1000

            # Handling meter resets and anomalies
            df["KWH_DIFF"] = df["KWH_DIFF"].apply(lambda x: x if x >= 0 else 0)  # Remove negative resets
            df["KWH_DIFF"] = df["KWH_DIFF"].apply(lambda x: x if x < 1000 else 0)  # Remove extreme spikes

            df.rename(columns={"KWH_DIFF": "KWH"}, inplace=True)

            df.drop(columns=["KWH_VALUES"], inplace=True)

            # Convert to daily format
            df = df.resample("D").sum().fillna(0)

            # Convert to horizontal format
            df_transposed = df.T
            df_transposed.insert(0, "Meter Name", meter_name)

            category_data.append(df_transposed)

        if category_data:
            category_df = pd.concat(category_data)

            # Calculate total but **exclude specified meters**
            included_meters = [m for m in meters if m not in exclude_from_total.get(category, [])]
            total_df = category_df[category_df["Meter Name"].isin(included_meters)].iloc[:, 1:].sum().to_frame().T
            total_df.insert(0, "Meter Name", f"Total {category}")

            category_df = pd.concat([category_df, total_df])

            all_data.append(category_df)

    final_df = pd.concat(all_data, ignore_index=True)
    return final_df

def save_to_excel(df, meter_categories, filename="dashboard.xlsx", sheet_name="Dashboard"):

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Styles
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Define category colors (light shades)
    category_colors = {
        "Admin": "FFDDC1",
        "DG Yard": "D6EAF8",
        "Electronic": "D5F5E3",
        "Total HT": "FADBD8",
        "Total Lightning": "FCF3CF",
        "Total Mechanical": "D7BDE2",
        "Total Solar": "A2D9CE",
        "Transformer": "F5CBA7",
        "Utility": "85C1E9",
        "Water": "F1948A",
    }

    # Reverse lookup: {meter_name: category}
    meter_to_category = {}
    for category, meters in meter_categories.items():
        for meter in meters:
            meter_to_category[meter] = category

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Apply formatting
    for row in range(1, ws.max_row + 1):
        meter_name = ws[f"A{row}"].value

        # Apply header styling (first row)
        if row == 1:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = font_bold
                cell.alignment = alignment_center
                cell.fill = header_fill
                cell.border = border_thin

        else:
            # Assign category colors
            category = meter_to_category.get(meter_name.replace("Total ", ""), None)
            if category and category in category_colors:
                fill_color = PatternFill(start_color=category_colors[category], end_color=category_colors[category], fill_type="solid")
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill_color

            # Bold for total rows
            if "Total" in meter_name:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).font = font_bold

        # Apply border to all cells
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = border_thin

    wb.save(filename)
    print(f"Excel file '{filename}' with worksheet '{sheet_name}' saved successfully!")



meter_categories = {
    "Admin": ["Admin Main meter"],
    "DG Yard": ["Incoming 1010 KVA DG-1", "Incoming 1010 KVA DG -2","Incoming 1010 KVA DG-3","Incoming 500 KVA DG-4","Incoming 500 KVA DG -5","Incoming 500 KVA DG -6"],
    "Electronic": ["ERSA Machine -1", "ERSA Machine -2", "K Nock Line 3", "Nox Line 1 Backend", "Nox Line 1 Laser",
                   "SMT UPS Supply", "M4A Line 1B", "M4A Line 1A", "Test end 1", "Test end 2", "M4A Line 2B",
                   "M4A Line 2A", "K Nock Line 1 & 2", "DLA Line 1", "DLA Line 2", "DLA Line 3", "DLA Line 4",
                   "Electronics AHU-1", "Electronics AHU-2", "Electronics AHU-3", "Electronics AHU-4",
                   "Electronics AHU-5", "Electronics AHU-6", "Electronics AHU-7", "Electronics UPS Incomer",
                   "Electronics UPS Outgoing - 1", "Electronics UPS Outgoing - 2", "Electronics UPS Outgoing - 3",
                   "Electronics UPS Outgoing - 4", "Technical Penthouse-L-1 Incomer main panel", "Electronics AHU main panel"],
    "Total HT": ["2 in 1 Incomer-1","2 in 1 Outgoing-1","2 in 1 Outgoing-2","5 in 2 Incomer-1","5 in 2 Incomer-2","5 in 2 Outgoing-1","5 in 2 Outgoing-2","5 in 2 Outgoing-3","5 in 2 Outgoing-4"],
    "Total Lightning": ["Panel Room- LDB-Mech,Pump Room, security, WRC AHU", "Panel Room-Aux DB-1","Panel Room-Mech & WH ELDB","BMS Room- MLDB ( Electronics DB )","BMS Room-Aux DB - 2","BMS Room ELDB"],
    "Total Mechanical": ["AUTO BPM-3", "AUTO BPM-5", "AUTO BPM-7", "AUTO BPM-8", "BPM-2", "Brazing Oven -1", "Brazing Oven -2",
                   "Brazing Oven-5", "CHARGING COOLING STATION", "CLWS-1(Mechanical)", "CLWS-2(Mechanical)", "Deoiling Oven-01",
                   "Deoiling Oven-02", "Deoiling Oven-03", "FINAL INSPECTION 1 TO 5", "FLEECE WELDING", "Future Line -11",
                   "Future Line -12", "Matix Auto BPM Line 7", "MATRIX L-5", "Matrix L-9, BPM 7", "MATRIX LINE 9",
                   "Matrix Line-1", "Matrix Line-10", "MATRIX LINE-11", "Matrix Line-2", "MATRIX LINE-3", "MATRIX LINE-4",
                   "Matrix-L-8", "MBPM LINE-1", "MBPM LINE-2", "MBPM LINE-3", "MBPM LINE-4", "MBPM LINE-5", "MCC AHU",
                   "Mech UPS", "Mechanical Sub Panel", "OLD MATRIX LINE", "Oven-6", "Oven-3", "Oven-7", "PE PUNCHING LINE-1",
                   "PE PUNCHING LINE-2", "POWDER ROOM-1", "POWDER ROOM-2", "PRE OXIDATION OVEN", "Sample_Shop", "Spare 25 f1",
                   "Spare-14f1", "Spare-15f1", "Technical Penthouse -L-2 Main meter", "Mechanical AHU-01", "Mechanical AHU-02",
                   "Mechanical AHU-03", "Mechanical AHU-04", "PCWP-Mechanical", "Transformer No. 2", "Transformer No. 3"],
    "Total Solar": ["Solar incomer 2000A"],
    "Transformer": ["APFC 400 KVAR Capacitor - 1","APFC 750 KVAR Capacitor- 4","APFC 800 KVAR Capacitor - 2","APFC 800 KVAR Capacitor - 3","Transformer No. 1","Transformer No. 2","Transformer No. 3","Transformer No. 4"],
    "Utility": [
        "STP Energy Meter", "COOLING TOWER-2", "CHILLER-2", "Air cool chiller Mechanical  (CH4)",
        "COOLING TOWER-1", "CHILLER-1", "Air cool chiller Electronics (CH-3)",
        "AIR COMPRESSOR COMP-1", "AIR COMPRESSOR COMP-2", "Pump Room main meter",
        "Air Compressor Main meter"],
    "Water": [
            "20 KL admin water", "30 KL water", "40 KL water",
            "MIDC Water Meter", "Level-2 20 KL Tank"
        ],

}
# Define meters to exclude from totals
exclude_from_total = {
    "Total HT": ["2 in 1 Outgoing-1","2 in 1 Outgoing-2","5 in 2 Incomer-1","5 in 2 Incomer-2","5 in 2 Outgoing-1","5 in 2 Outgoing-2","5 in 2 Outgoing-3","5 in 2 Outgoing-4"],
    "Electronic": ["ERSA Machine -1", "ERSA Machine -2", "K Nock Line 3", "Nox Line 1 Backend", "Nox Line 1 Laser",
                   "SMT UPS Supply", "M4A Line 1B", "M4A Line 1A", "Test end 1", "Test end 2", "M4A Line 2B",
                   "M4A Line 2A", "K Nock Line 1 & 2", "DLA Line 1", "DLA Line 2", "DLA Line 3", "DLA Line 4",
                   "Electronics AHU-1", "Electronics AHU-2", "Electronics AHU-3", "Electronics AHU-4",
                   "Electronics AHU-5", "Electronics AHU-6", "Electronics AHU-7", "Electronics UPS Incomer",
                   "Electronics UPS Outgoing - 1", "Electronics UPS Outgoing - 2", "Electronics UPS Outgoing - 3",
                   "Electronics UPS Outgoing - 4"],
    "Total Lightning": ["BMS Room ELDB"],
    "Total Mechanical": ["AUTO BPM-3", "AUTO BPM-5", "AUTO BPM-7", "AUTO BPM-8", "BPM-2", "Brazing Oven -1", "Brazing Oven -2",
                   "Brazing Oven-5", "CHARGING COOLING STATION", "CLWS-1(Mechanical)", "CLWS-2(Mechanical)", "Deoiling Oven-01",
                   "Deoiling Oven-02", "Deoiling Oven-03", "FINAL INSPECTION 1 TO 5", "FLEECE WELDING", "Future Line -11",
                   "Future Line -12", "Matix Auto BPM Line 7", "MATRIX L-5", "Matrix L-9, BPM 7", "MATRIX LINE 9",
                   "Matrix Line-1", "Matrix Line-10", "MATRIX LINE-11", "Matrix Line-2", "MATRIX LINE-3", "MATRIX LINE-4",
                   "Matrix-L-8", "MBPM LINE-1", "MBPM LINE-2", "MBPM LINE-3", "MBPM LINE-4", "MBPM LINE-5", "MCC AHU",
                   "Mech UPS", "Mechanical Sub Panel", "OLD MATRIX LINE", "Oven-6", "Oven-3", "Oven-7", "PE PUNCHING LINE-1",
                   "PE PUNCHING LINE-2", "POWDER ROOM-1", "POWDER ROOM-2", "PRE OXIDATION OVEN", "Sample_Shop", "Spare 25 f1",
                   "Spare-14f1", "Spare-15f1", "Technical Penthouse -L-2 Main meter"],
    "Transformer": ["APFC 400 KVAR Capacitor - 1","APFC 750 KVAR Capacitor- 4","APFC 800 KVAR Capacitor - 2","APFC 800 KVAR Capacitor - 3"],
}

def fetch_all_meters_dataframe(start_time, end_time, interval="30min", values_to_fetch=["kwh", "pf", "va", "current"]):

    # MongoDB Connection
    client = pymongo.MongoClient("mongodb://192.168.0.34:27017/")
    db = client["edms"]
    collections = {
        "kwh": db["kwh"],
        "pf": db["pf"],
        "va": db["va"],
        "current": db["current"]
    }

    def fetch_data(collection, field_name):
        cursor = collection.find({
            "timestamp": {"$gte": start_time.strftime("%Y-%m-%d %H:%M:%S"),
                          "$lt": end_time.strftime("%Y-%m-%d %H:%M:%S")}
        })

        data = {}

        for doc in cursor:
            timestamp = pd.to_datetime(doc["timestamp"])
            for i, meter in enumerate(doc["device_names"]):
                if meter not in data:
                    data[meter] = {"Timestamp": [], field_name.upper(): []}
                data[meter]["Timestamp"].append(timestamp)
                data[meter][field_name.upper()].append(doc[field_name][i])

        dataframes = {}
        for meter, meter_data in data.items():
            df = pd.DataFrame(meter_data).sort_values(by="Timestamp").set_index("Timestamp")
            dataframes[meter] = df

        return dataframes

    meter_dataframes = {}

    for value in values_to_fetch:
        if value in collections:
            fetched_data = fetch_data(collections[value], f"{value}_values")
            for meter, df in fetched_data.items():
                if meter not in meter_dataframes:
                    meter_dataframes[meter] = df
                else:
                    meter_dataframes[meter] = meter_dataframes[meter].join(df, how="outer")

    # List of meters that need KWH values divided by 1000
    meters_to_divide = {
        "Pump Room main meter",
        "Mechanical AHU-01",
        "Mechanical AHU-02",
        "Mechanical AHU-03",
        "Mechanical AHU-04",
        "Technical Penthouse -L-2 Main meter",
        "Electronics AHU-1",
        "Electronics AHU-2",
        "Electronics AHU-4",
        "Electronics AHU-5",
        "Electronics AHU-6",
        "Electronics AHU-7",
        "PCWP-Mechanical",
        "Air cool chiller Mechanical (CH4)",
        "Air cool chiller Electronics (CH-3)",
        "CLWS-2 (Mechanical)",
        "Air Compressor Main meter",
        "Technical Penthouse-L-1 Incomer main panel",
        "Electronics UPS Incomer",
        "Electronics UPS Outgoing - 1",
        "Electronics UPS Outgoing - 3",
        "BMS Room ELDB",
        "BMS Room-Aux DB - 2",
        "BMS Room- MLDB (Electronics DB)",
        "Electronics AHU main panel",
        "Canteen main panel",
        "Transformer No. 1",
        "Transformer No. 2",
        "Transformer No. 3",
        "Transformer No. 4",
        "Incoming 500 KVA DG -6",
        "Panel Room-Mech & WH ELDB",
        "Panel Room-Aux DB-1","Panel Room- LDB-Mech,Pump Room, security, WRC AHU","BMS Room- MLDB ( Electronics DB )"
    }

    # Compute kWh difference safely
    for meter, df in meter_dataframes.items():
        if "KWH_VALUES" in df.columns:
            df["KWH_DIFF"] = df["KWH_VALUES"].diff().fillna(0)

            # Handling meter resets and anomalies
            df["KWH_DIFF"] = df["KWH_DIFF"].apply(lambda x: x if x >= 0 else 0)  # Remove negative resets
            df["KWH_DIFF"] = df["KWH_DIFF"].apply(lambda x: x if x < 1000 else 0)  # Remove extreme spikes

            df.rename(columns={"KWH_DIFF": "KWH"}, inplace=True)
            df.drop(columns=["KWH_VALUES"], inplace=True)

            # If the meter is in the list, divide KWH values by 1000
            if meter in meters_to_divide:
                df["KWH"] /= 1000

        if "CURRENT_VALUES" in df.columns:
            df.rename(columns={"CURRENT_VALUES": "KW/FLOW RATE"}, inplace=True)

    # Resample based on the selected interval

    return meter_dataframes


def saave_to_excel(meter_dataframes, output_filename="energy_meter_data.xlsx"):
    with pd.ExcelWriter(output_filename, engine="xlsxwriter") as writer:
        workbook = writer.book

        for meter, df in meter_dataframes.items():
            sheet_name = meter[:31]  # Excel sheet names must be <= 31 characters
            worksheet = workbook.add_worksheet(sheet_name)
            writer.sheets[sheet_name] = worksheet  # Register sheet manually

            df.replace([float("inf"), float("-inf")], 0, inplace=True)
            df.fillna(0, inplace=True)  # Replace missing values with 0 instead of empty string

            # Convert timestamp column if necessary
            if not pd.api.types.is_datetime64_any_dtype(df.index):
                df.index = pd.to_datetime(df.index)

            # Ensure numeric columns are actually numeric
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")  # Convert strings to NaN, then fill with 0
            df.fillna(0, inplace=True)  # Replace NaN with 0

            # Resample data to 30-minute intervals by summing every 3 rows
            df_resampled = df.resample("30min").sum()

            # Formatting
            title_format = workbook.add_format({"bold": True, "font_size": 14, "align": "center"})
            header_format = workbook.add_format({"bold": True, "bg_color": "#DDEBF7", "border": 1})
            cell_format = workbook.add_format({"border": 1})

            # Merge and write meter name at the top
            worksheet.merge_range("A1:E1", meter, title_format)

            # Write column headers manually
            headers = ["Timestamp"] + list(df_resampled.columns)
            for col_num, value in enumerate(headers):
                worksheet.write(2, col_num, value, header_format)

            # Write DataFrame contents
            for row_num, (timestamp, row) in enumerate(df_resampled.iterrows(), start=3):
                worksheet.write(row_num, 0, timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                for col_num, value in enumerate(row, start=1):
                    worksheet.write(row_num, col_num, value, cell_format)

            worksheet.set_column(0, 0, 22)
            for i, col in enumerate(df_resampled.columns, start=1):
                worksheet.set_column(i, i, max(len(col), 10) + 2)

            # Add bar chart for kWh
            if "KWH" in df_resampled.columns:
                chart = workbook.add_chart({"type": "column"})
                chart.add_series({
                    "name": f"{meter} - kWh Usage",
                    "categories": [sheet_name, 3, 0, len(df_resampled) + 3, 0],
                    "values": [sheet_name, 3, df_resampled.columns.get_loc("KWH") + 1, len(df_resampled) + 3, df_resampled.columns.get_loc("KWH") + 1],
                })
                chart.set_title({"name": "30-Minute kWh Usage"})
                chart.set_x_axis({"name": "Time"})
                chart.set_y_axis({"name": "kWh"})
                chart.set_size({"width": 600, "height": 500})
                worksheet.insert_chart(3, len(df_resampled.columns) + 2, chart)

                # Compute daily total kWh
                daily_totals = df_resampled["KWH"].resample("D").sum()
                daily_totals_df = daily_totals.reset_index()
                daily_totals_df.columns = ["Date", "Total kWh"]

                # Write daily total table at Q3
                start_col = 16  # Column Q (0-based index)
                start_row = 2  # Row 3 (0-based index)
                worksheet.set_column(16, 17, 20)
                worksheet.write(start_row, start_col, "Daily Total Consumption", title_format)

                for col_num, column_name in enumerate(daily_totals_df.columns):
                    worksheet.write(start_row + 1, start_col + col_num, column_name, header_format)

                for row_num, (date, total) in enumerate(daily_totals_df.itertuples(index=False), start=start_row + 2):
                    worksheet.write(row_num, start_col, date.strftime("%Y-%m-%d"))
                    worksheet.write_number(row_num, start_col + 1, total)

    print(f"Excel file '{output_filename}' saved successfully.")

current_date = datetime.now()

start_time = datetime(current_date.year, current_date.month, 1)

next_month = current_date.month % 12 + 1
next_month_year = current_date.year if next_month > 1 else current_date.year + 1
end_time = datetime(next_month_year, next_month, 1)

print("Start Time:", start_time)
print("End Time:", end_time)


import os
import datetime
import smtplib
import pandas as pd
from email.message import EmailMessage


# Fetch data and save to Excel
df = fetch_monthly_kwh_for_meters(start_time, end_time, meter_categories, exclude_from_total)
meter_dataframes = fetch_all_meters_dataframe(start_time, end_time, interval="30min")  # Fix "H" to "h"

# Generate filenames with today's date
today_date = datetime.datetime.today().strftime("%Y-%m-%d")
dashboard_filename = f"dashboard_{today_date}.xlsx"
meter_filename = f"energy_meter_data_{today_date}.xlsx"


save_to_excel(df,meter_categories, dashboard_filename)
saave_to_excel(meter_dataframes, meter_filename)

from datetime import datetime
current_year = datetime.now().year
current_month = datetime.now().month
current_day = datetime .now().day

maintenance_filename = f"PM_MAINTENANCE_{current_day}_{current_month}_{current_year}.xlsx"
module_filename = f"PM_Module_{current_day}_{current_month}_{current_year}.xlsx"

# Email details
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "webinn15@gmail.com"
SENDER_PASSWORD = "rphz cgtg rwnv ojgn"
RECEIVER_EMAIL = "instatushir@gmail.com"

# Create email
msg = EmailMessage()
msg["Subject"] = f"Energy Meter & Maintenance Reports - {today_date}"
msg["From"] = SENDER_EMAIL
msg["To"] = RECEIVER_EMAIL
msg.set_content(f"Please find attached energy meter reports and pm reports for {today_date}.")

# Attach Excel files
for file in [dashboard_filename, meter_filename, maintenance_filename, module_filename]:
    with open(file, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=file)

# Send email
try:
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
    print("Email sent successfully.")

    # Remove files after sending
    os.remove(dashboard_filename)
    os.remove(meter_filename)
    os.remove(maintenance_filename)
    os.remove(module_filename)
    print("Files deleted.")
except Exception as e:
    print(f"Error: {e}")
